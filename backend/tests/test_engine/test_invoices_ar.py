"""Tests for the deterministic open AR -> Light Invoices (AR) engine.

The fixture replicates the real "AR example Sunshine.xlsx": a QuickBooks-style
A/R Aging Detail report grouped by customer, plus a filled-in Invoices (AR)
template sheet whose single row is the worked example for INV360.
"""

from datetime import datetime

import pytest

from app.engine.invoices_ar import (
    LIGHT_INVOICES_AR_COLUMNS,
    analyze_invoices_ar_workbook,
    apply_structured_updates_to_analysis,
    format_invoices_ar_analysis_message,
    transform_open_ar_to_light_invoices,
)

_HEADER = [
    "Customer",
    "Transaction Type",
    "Date",
    "Document Number",
    "P.O. No.",
    "Due Date",
    "Age",
    "Open Balance",
]

_TEMPLATE_HEADER = LIGHT_INVOICES_AR_COLUMNS

# (customer group label, [(txn type, date, doc no, po, due date, age, amount)])
_DEFAULT_GROUPS = [
    (
        "C117 Syneos Health, LLC",
        [
            ("Invoice", datetime(2026, 5, 12), "INV360", "PO_2001071695_1", datetime(2026, 7, 11), 33, 375000.0),
            ("Invoice", datetime(2026, 7, 24), "INV364", "PO_2001071695_1", datetime(2026, 8, 23), -10, 375000.0),
        ],
    ),
    (
        "C124 Omnicom Advertising",
        [
            ("Invoice", datetime(2026, 5, 12), "INV359", "Purchase Order 4500006667", datetime(2026, 7, 11), 33, 130650.0),
        ],
    ),
    (
        "C141 Jackson Lewis P.C.",
        [
            ("Invoice", datetime(2026, 5, 12), "INV358", None, datetime(2026, 7, 11), 33, 10777.78),
            ("Invoice", datetime(2026, 6, 29), "INV363", None, datetime(2026, 8, 28), -15, 22222.22),
        ],
    ),
]

_EXAMPLE_ROW = {
    "Entity": "causaLens US",
    "Invoice Status": "Open",
    "Customer": "Syneos Health, LLC",
    "Invoice Number": "INV360",
    "Invoice Date": datetime(2026, 5, 12),
    "Due Date": datetime(2026, 7, 11),
    "Currency": "USD",
    "Description": "Data migration - INV360",
    "Invoice Currency Amount": 375000.0,
    "Local Currency Amount": 375000.0,
    "Product": "Data migration",
    "Quantity": 1.0,
    "Unit Price": 375000.0,
}


def _ar_workbook(path, *, with_template=True, groups=_DEFAULT_GROUPS):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR US"

    ws.append(["causaLens"])
    ws.append(["causaLens : causaLens US"])
    ws.append(["A/R Aging Detail"])
    ws.append(["As of 31 July 2026"])
    ws.append([None])
    ws.append([None])
    ws.append(_HEADER)

    grand_total = 0.0
    for label, rows in groups:
        ws.append([label])
        subtotal = 0.0
        for txn_type, date, doc, po, due, age, amount in rows:
            ws.append([None, txn_type, date, doc, po, due, age, amount])
            subtotal += amount
        ws.append([f"Total - {label}", None, None, None, None, None, None, round(subtotal, 2)])
        grand_total += subtotal
    ws.append(["Total", None, None, None, None, None, None, round(grand_total, 2)])

    if with_template:
        template = wb.create_sheet("Invoices (AR)")
        template.append(_TEMPLATE_HEADER)
        template.append([_EXAMPLE_ROW.get(col) for col in _TEMPLATE_HEADER])

    wb.save(path)
    return path


def _row_by_invoice(df, invoice_number):
    matches = df[df["Invoice Number"] == invoice_number]
    assert len(matches) == 1, f"expected exactly one row for {invoice_number}"
    return matches.iloc[0]


def test_detects_grouped_report_template_and_facts(tmp_path):
    path = _ar_workbook(tmp_path / "AR example.xlsx")
    analysis = analyze_invoices_ar_workbook(path)

    assert analysis.source_sheet == "AR US"
    assert analysis.template_sheet == "Invoices (AR)"
    assert analysis.layout == "grouped"
    assert analysis.roles["customer_name"] == "Customer"
    assert analysis.roles["invoice_number"] == "Document Number"
    assert analysis.roles["po_number"] == "P.O. No."
    assert analysis.roles["open_amount"] == "Open Balance"

    # Facts come from the worked example, not the title block.
    assert analysis.facts["entity"] == "causaLens US"
    assert analysis.facts["currency"] == "USD"
    assert analysis.facts["description_prefix"] == "Data migration"
    assert analysis.facts["product"] == "Data migration"

    assert analysis.detail_rows == 5
    assert analysis.invoice_rows == 5
    assert analysis.invoice_total == pytest.approx(913650.0)
    assert analysis.reported_total == pytest.approx(913650.0)
    assert analysis.source_total == pytest.approx(913650.0)
    assert analysis.ready
    assert not analysis.questions


def test_transform_matches_worked_example(tmp_path):
    path = _ar_workbook(tmp_path / "AR example.xlsx")
    analysis = analyze_invoices_ar_workbook(path)
    df = transform_open_ar_to_light_invoices(path, analysis)

    assert list(df.columns) == LIGHT_INVOICES_AR_COLUMNS
    assert len(df) == 5
    assert df["Invoice Currency Amount"].sum() == pytest.approx(913650.0)
    assert set(df["Invoice Status"]) == {"Open"}

    row = _row_by_invoice(df, "INV360")
    assert row["Entity"] == "causaLens US"
    assert row["Customer"] == "Syneos Health, LLC"
    assert row["Customer ID"] == "C117"
    assert row["Invoice Date"] == "2026-05-12"
    assert row["Due Date"] == "2026-07-11"
    assert row["Currency"] == "USD"
    assert row["PO Number"] == "PO_2001071695_1"
    assert row["Description"] == "Data migration - INV360"
    assert row["Invoice Currency Amount"] == pytest.approx(375000.0)
    assert row["Local Currency Amount"] == pytest.approx(375000.0)
    assert row["Product"] == "Data migration"
    assert row["Quantity"] == 1
    assert row["Unit Price"] == pytest.approx(375000.0)
    assert row["Tax Code"] is None
    assert row["Account"] is None

    other = _row_by_invoice(df, "INV359")
    assert other["Customer"] == "Omnicom Advertising"
    assert other["Customer ID"] == "C124"

    no_po = _row_by_invoice(df, "INV358")
    assert no_po["PO Number"] is None


def test_subtotals_and_grand_total_are_never_invoices(tmp_path):
    path = _ar_workbook(tmp_path / "AR example.xlsx")
    analysis = analyze_invoices_ar_workbook(path)
    df = transform_open_ar_to_light_invoices(path, analysis)

    assert not any(str(c).lower().startswith("total") for c in df["Customer"])
    # Grand total is kept as the reconciliation anchor instead.
    assert analysis.reported_total == pytest.approx(913650.0)


def test_payments_become_credit_notes_not_netted(tmp_path):
    groups = [
        (
            "C117 Syneos Health, LLC",
            _DEFAULT_GROUPS[0][1]
            + [("Payment", datetime(2026, 7, 30), "PMT-1", None, None, None, -100000.0)],
        ),
        _DEFAULT_GROUPS[1],
        _DEFAULT_GROUPS[2],
    ]
    path = _ar_workbook(tmp_path / "AR with payment.xlsx", groups=groups)
    analysis = analyze_invoices_ar_workbook(path)

    # Nothing is netted: invoices keep their gross open balance and the
    # payment becomes a credit note with a negative amount.
    assert analysis.invoice_rows == 5
    assert analysis.credit_note_rows == 1
    assert analysis.invoice_total == pytest.approx(913650.0)
    assert analysis.credit_note_total == pytest.approx(-100000.0)
    assert analysis.source_total == pytest.approx(813650.0)
    assert analysis.reported_total == pytest.approx(813650.0)

    df = transform_open_ar_to_light_invoices(path, analysis)
    assert len(df) == 6
    assert df["Invoice Currency Amount"].sum() == pytest.approx(813650.0)

    oldest = _row_by_invoice(df, "INV360")
    assert oldest["Invoice Currency Amount"] == pytest.approx(375000.0)

    credit_note = _row_by_invoice(df, "PMT-1")
    assert credit_note["Invoice Currency Amount"] == pytest.approx(-100000.0)
    assert credit_note["Unit Price"] == pytest.approx(-100000.0)
    assert credit_note["Customer"] == "Syneos Health, LLC"

    message = format_invoices_ar_analysis_message(analysis)
    assert "credit note" in message.lower()


def test_customer_in_credit_is_emitted_as_negative_credit_note(tmp_path):
    groups = _DEFAULT_GROUPS + [
        (
            "C999 Overpaid Ltd",
            [("Payment", datetime(2026, 7, 1), "PMT-9", None, None, None, -500.0)],
        )
    ]
    path = _ar_workbook(tmp_path / "AR with credit.xlsx", groups=groups)
    analysis = analyze_invoices_ar_workbook(path)

    assert [c["customer_name"] for c in analysis.credit_note_customers] == ["Overpaid Ltd"]
    assert analysis.credit_note_total == pytest.approx(-500.0)

    df = transform_open_ar_to_light_invoices(path, analysis)
    assert len(df) == 6
    credit_note = _row_by_invoice(df, "PMT-9")
    assert credit_note["Customer"] == "Overpaid Ltd"
    assert credit_note["Invoice Currency Amount"] == pytest.approx(-500.0)


def test_settled_customer_keeps_both_open_items(tmp_path):
    groups = _DEFAULT_GROUPS + [
        (
            "C888 Square Ltd",
            [
                ("Invoice", datetime(2026, 6, 1), "INV900", None, datetime(2026, 7, 1), 10, 200.0),
                ("Payment", datetime(2026, 7, 1), "PMT-8", None, None, None, -200.0),
            ],
        )
    ]
    path = _ar_workbook(tmp_path / "AR settled.xlsx", groups=groups)
    analysis = analyze_invoices_ar_workbook(path)

    # The aging report still shows both items open, so both are migrated;
    # they net to zero inside Light once the credit is applied there.
    df = transform_open_ar_to_light_invoices(path, analysis)
    assert len(df) == 7
    assert _row_by_invoice(df, "INV900")["Invoice Currency Amount"] == pytest.approx(200.0)
    assert _row_by_invoice(df, "PMT-8")["Invoice Currency Amount"] == pytest.approx(-200.0)
    assert df["Invoice Currency Amount"].sum() == pytest.approx(913650.0)


def test_single_customer_report_is_still_grouped(tmp_path):
    """Regression: "AR cL UK.xlsx" — one customer, three invoices.

    A single group header used to fail the grouped-layout threshold, the sheet
    was read flat, every blank-customer detail row was dropped, and the user
    was told everything was settled. One customer must be enough.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARAgingDetail"
    ws.append(["causaLens"])
    ws.append(["causaLens"])
    ws.append(["A/R Aging Detail"])
    ws.append(["As of 31 July 2026"])
    ws.append([None])
    ws.append([None])
    ws.append(_HEADER)
    ws.append(["C010 Cisco Systems Inc."])
    ws.append([None, "Invoice", datetime(2026, 3, 11), "INV346", "PO: USA000BL1016227CW", datetime(2026, 5, 10), 95, 7432.49])
    ws.append([None, "Invoice", datetime(2026, 5, 1), "INV347", "PO USA000BL1016227CW", datetime(2026, 6, 30), 44, 7432.49])
    ws.append([None, "Invoice", datetime(2026, 5, 29), "INV348", "PO USA000BL1016227CW", datetime(2026, 7, 28), 16, 7432.49])
    ws.append(["Total - C010 Cisco Systems Inc.", None, None, None, None, None, None, 22297.47])
    ws.append(["Total", None, None, None, None, None, None, 22297.47])
    path = tmp_path / "AR cL UK.xlsx"
    wb.save(path)

    analysis = analyze_invoices_ar_workbook(path)
    assert analysis.source_sheet == "ARAgingDetail"
    assert analysis.layout == "grouped"
    assert analysis.detail_rows == 3
    assert analysis.invoice_rows == 3
    assert analysis.credit_note_rows == 0
    assert analysis.source_total == pytest.approx(22297.47)
    assert analysis.reported_total == pytest.approx(22297.47)
    # Entity comes from the title block; only currency should be missing.
    assert analysis.facts["entity"] == "causaLens"
    assert [q for q in analysis.questions if "currency" in q.lower()]

    updated = apply_structured_updates_to_analysis(analysis, {"currency": "GBP"}, path)
    assert updated.ready

    df = transform_open_ar_to_light_invoices(path, updated)
    assert len(df) == 3
    assert df["Invoice Currency Amount"].sum() == pytest.approx(22297.47)
    row = _row_by_invoice(df, "INV346")
    assert row["Customer"] == "Cisco Systems Inc."
    assert row["Customer ID"] == "C010"
    assert row["Currency"] == "GBP"


def test_entity_comes_from_colon_title_without_template(tmp_path):
    path = _ar_workbook(tmp_path / "AR no template.xlsx", with_template=False)
    analysis = analyze_invoices_ar_workbook(path)

    # QuickBooks "company : entity" titles resolve to the leaf entity.
    assert analysis.facts["entity"] == "causaLens US"
    # The report has no currency column and no worked example to take it from.
    assert not analysis.facts.get("currency")
    assert not analysis.ready
    assert any("currency" in q.lower() for q in analysis.questions)


def test_structured_updates_toggle_split_and_po(tmp_path):
    path = _ar_workbook(tmp_path / "AR example.xlsx")
    analysis = analyze_invoices_ar_workbook(path)

    updated = apply_structured_updates_to_analysis(
        analysis,
        {"split_customer_code": "false", "include_po_number": "false", "currency": "GBP"},
        path,
    )
    assert updated.facts["currency"] == "GBP"
    assert updated.facts["split_customer_code"] is False

    df = transform_open_ar_to_light_invoices(path, updated)
    row = _row_by_invoice(df, "INV360")
    assert row["Customer"] == "C117 Syneos Health, LLC"
    assert row["Customer ID"] is None
    assert row["PO Number"] is None
    assert row["Currency"] == "GBP"


def test_rejects_unknown_structured_update(tmp_path):
    path = _ar_workbook(tmp_path / "AR example.xlsx")
    analysis = analyze_invoices_ar_workbook(path)

    with pytest.raises(ValueError, match="Unsupported invoices AR updates"):
        apply_structured_updates_to_analysis(analysis, {"ap_account": "221101"}, path)
