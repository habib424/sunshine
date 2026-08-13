import re
import uuid
from pathlib import Path

import openpyxl
import pandas as pd

from app.config import settings

# Matches formulas that are just a numeric literal, e.g. "=91778.72" or "=(5)".
_CONST_FORMULA_RE = re.compile(r"^=\s*\(?\s*([+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)\s*\)?\s*$")


def ensure_readable_workbook(file_path: Path) -> Path:
    """Return a path whose cell values are trustworthy for reading.

    Some ERP exports store amounts as constant formulas (e.g. "=91778.72")
    and save the workbook without recalculating, leaving cached values at 0
    or empty. pandas/openpyxl read cached values, so every amount appears
    blank or zero. When that happens, write a sibling
    "<name>.computed.xlsx" with the constant formulas materialized and
    return its path. The original upload is never modified.
    """
    if file_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return file_path

    computed = file_path.with_name(f"{file_path.stem}.computed{file_path.suffix}")
    if computed.exists() and computed.stat().st_mtime >= file_path.stat().st_mtime:
        return computed

    if not _has_stale_constant_formulas(file_path):
        return file_path

    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return file_path

    for ws in wb_formulas.worksheets:
        ws_values = wb_values[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                cached = ws_values.cell(row=cell.row, column=cell.column).value
                match = _CONST_FORMULA_RE.match(cell.value)
                if match:
                    cell.value = float(match.group(1))
                else:
                    # Non-constant formula: best available value is the cache.
                    cell.value = cached

    wb_formulas.save(computed)
    return computed


def _has_stale_constant_formulas(file_path: Path) -> bool:
    """Fast streaming check: does any "=<number>" formula disagree with its cache?

    Uses two read-only passes (cheap) so the common case — workbooks with no
    formulas or with fresh caches — never pays for the full rewrite.
    """
    try:
        constants: dict[str, dict[tuple[int, int], float]] = {}
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        for ws in wb.worksheets:
            per_sheet: dict[tuple[int, int], float] = {}
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, value in enumerate(row, start=1):
                    if isinstance(value, str) and value.startswith("="):
                        match = _CONST_FORMULA_RE.match(value)
                        if match:
                            per_sheet[(row_idx, col_idx)] = float(match.group(1))
            if per_sheet:
                constants[ws.title] = per_sheet
        wb.close()

        if not constants:
            return False

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            for title, per_sheet in constants.items():
                ws = wb[title]
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    for col_idx, cached in enumerate(row, start=1):
                        literal = per_sheet.get((row_idx, col_idx))
                        if literal is None:
                            continue
                        if not isinstance(cached, (int, float)) or abs(float(cached) - literal) > 1e-9:
                            return True
            return False
        finally:
            wb.close()
    except Exception:
        return False


def delete_upload_files(file_path: Path) -> None:
    """Delete an upload from disk, including its original/.computed sibling."""
    candidates = {file_path}
    if file_path.stem.endswith(".computed"):
        candidates.add(file_path.with_name(file_path.name.replace(".computed", "", 1)))
    else:
        candidates.add(file_path.with_name(f"{file_path.stem}.computed{file_path.suffix}"))
    for candidate in candidates:
        if candidate.exists():
            candidate.unlink()


def save_upload(file_bytes: bytes, original_name: str) -> tuple[str, Path]:
    file_id = str(uuid.uuid4())
    ext = Path(original_name).suffix
    filename = f"{file_id}{ext}"
    dest = settings.uploads_path / filename
    dest.write_bytes(file_bytes)
    # If amounts are stored as uncalculated formulas, materialize a readable
    # copy and make that the canonical path; the original bytes stay on disk.
    dest = ensure_readable_workbook(dest)
    return file_id, dest


def read_file_metadata(file_path: Path) -> dict:
    ext = file_path.suffix.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        return _read_excel_metadata(file_path)
    elif ext == ".csv":
        return _read_csv_metadata(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _read_excel_metadata(file_path: Path) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None:
            headers.append(str(cell.value))

    row_count = ws.max_row - 1 if ws.max_row else 0  # exclude header
    wb.close()

    return {
        "sheet_names": sheet_names,
        "column_headers": headers,
        "row_count": row_count,
    }


def _read_csv_metadata(file_path: Path) -> dict:
    df = pd.read_csv(file_path, nrows=0)
    row_count = sum(1 for _ in open(file_path)) - 1
    return {
        "sheet_names": ["Sheet1"],
        "column_headers": list(df.columns),
        "row_count": max(row_count, 0),
    }


def read_file_preview(file_path: Path, limit: int = 20, sheet: int | str = 0, header_row: int = 0) -> dict:
    ext = file_path.suffix.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        df = pd.read_excel(file_path, sheet_name=sheet, header=header_row, nrows=limit, engine="openpyxl")
    elif ext == ".csv":
        df = pd.read_csv(file_path, header=header_row, nrows=limit)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    df = df.fillna("")
    return {
        "headers": list(df.columns.astype(str)),
        "rows": df.values.tolist(),
        "total_rows": len(df),
    }


def read_file_as_dataframe(file_path: Path, sheet: int | str = 0, header_row: int = 0, skip_footer: int = 0) -> pd.DataFrame:
    ext = file_path.suffix.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        df = pd.read_excel(file_path, sheet_name=sheet, header=header_row, engine="openpyxl")
    elif ext == ".csv":
        df = pd.read_csv(file_path, header=header_row)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    if skip_footer > 0:
        df = df.iloc[:-skip_footer]

    return df
