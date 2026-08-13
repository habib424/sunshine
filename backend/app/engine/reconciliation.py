from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


_NON_ALNUM = re.compile(r"[^a-z0-9]+")
_ACCOUNT_AT_START = re.compile(r"^\s*(\d+)(?:\.0+)?(?:\s+|$)")
_INVALID_SHEET_CHARS = set(r':\/?*[]')


def _normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return _NON_ALNUM.sub(" ", text.lower()).strip()


def _compact(value: Any) -> str:
    return _normalize(value).replace(" ", "")


def _read_table(path: Path, required_headers: list[str] | None = None) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == ".csv":
        return pd.read_csv(path)

    if ext not in (".xlsx", ".xls", ".xlsm"):
        raise ValueError(f"Unsupported reconciliation file format: {ext}")

    engine = "openpyxl" if ext in (".xlsx", ".xlsm") else None
    raw = pd.read_excel(path, sheet_name=0, header=None, engine=engine)
    header_row = 0
    if required_headers:
        wanted = {_normalize(h) for h in required_headers}
        best_score = -1
        for idx, row in raw.head(80).iterrows():
            seen = {_normalize(v) for v in row.tolist() if pd.notna(v)}
            score = len(wanted & seen)
            if score > best_score:
                best_score = score
                header_row = int(idx)
            if score == len(wanted):
                break

    df = pd.read_excel(path, sheet_name=0, header=header_row, engine=engine)
    return df.dropna(how="all")


def _find_column(df: pd.DataFrame, aliases: list[str], *, contains: bool = False) -> str:
    normalized = {_normalize(col): col for col in df.columns}
    for alias in aliases:
        key = _normalize(alias)
        if key in normalized:
            return normalized[key]

    if contains:
        for alias in aliases:
            key = _normalize(alias)
            for col in df.columns:
                col_key = _normalize(col)
                if key and (key in col_key or col_key in key):
                    return col

    raise ValueError(f"Could not find any of these columns: {', '.join(aliases)}")


def _account_code(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    match = _ACCOUNT_AT_START.match(text)
    if match:
        return match.group(1)
    return text


def _account_description(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    match = _ACCOUNT_AT_START.match(text)
    if match:
        return text[match.end():].strip()
    return text


def _amount(value: Any) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, int | float):
        return float(value)

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return 0.0

    is_negative = False
    if text.startswith("(") and text.endswith(")"):
        is_negative = True
        text = text[1:-1]
    if text.endswith("-"):
        is_negative = True
        text = text[:-1]

    text = (
        text.replace(",", "")
        .replace("\u00a0", "")
        .replace("€", "")
        .replace("$", "")
        .replace("£", "")
        .strip()
    )
    if not text:
        return 0.0
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if is_negative else number


def _read_light_journal(path: Path) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame]:
    df = _read_table(path, ["Entity", "Account code", "Debit", "Credit"])

    entity_col = _find_column(df, ["Entity", "Company", "Legal entity"])
    account_col = _find_column(df, ["Account code", "Account", "GL account", "Ledger account"], contains=True)
    debit_col = _find_column(df, ["Debit", "Debit amount"], contains=True)
    credit_col = _find_column(df, ["Credit", "Credit amount"], contains=True)
    currency_col = None
    entry_col = None
    try:
        currency_col = _find_column(df, ["Currency", "Currency code"], contains=True)
    except ValueError:
        pass
    try:
        entry_col = _find_column(df, ["Entry Id", "Entry ID", "Journal entry id"], contains=True)
    except ValueError:
        pass

    working = pd.DataFrame(index=df.index)
    working["source_file"] = path.name
    working["entity"] = df[entity_col].astype(str).str.strip()
    working["account_code"] = df[account_col].apply(_account_code)
    working["entry_id"] = df[entry_col].astype(str).str.strip() if entry_col else ""
    working["currency"] = df[currency_col].astype(str).str.strip() if currency_col else ""
    working["debit"] = df[debit_col].apply(_amount)
    working["credit"] = df[credit_col].apply(_amount)
    working = working[(working["entity"] != "") & (working["account_code"] != "")]
    working["light_balance"] = working["debit"] - working["credit"]
    source_indexes = list(working.index)
    working = working.reset_index(drop=True)
    for col_idx, col_name in enumerate(df.columns):
        export_name = f"Original - {col_name}"
        working[export_name] = df.iloc[source_indexes, col_idx].reset_index(drop=True)

    grouped = (
        working.groupby(["entity", "account_code"], dropna=False)
        .agg(
            light_debit=("debit", "sum"),
            light_credit=("credit", "sum"),
            light_balance=("light_balance", "sum"),
            light_lines=("light_balance", "size"),
            light_entries=("entry_id", lambda s: int(s.replace("", pd.NA).dropna().nunique())),
            currency=("currency", lambda s: ", ".join(sorted({v for v in s if v and v.lower() != "nan"}))),
        )
        .reset_index()
    )

    meta = {
        "rows": int(len(working)),
        "entities": sorted(working["entity"].dropna().unique().tolist()),
        "accounts": int(grouped["account_code"].nunique()),
        "total_debit": float(working["debit"].sum()),
        "total_credit": float(working["credit"].sum()),
    }
    return grouped, meta, working


def _balance_column(df: pd.DataFrame) -> str:
    candidates = list(df.columns)
    for col in candidates:
        key = _normalize(col)
        if key.startswith("balance") and "account currency" not in key:
            return col
    for col in candidates:
        if _normalize(col).startswith("balance"):
            return col
    numeric_counts = [(col, pd.to_numeric(df[col].map(_amount), errors="coerce").notna().sum()) for col in candidates]
    numeric_counts.sort(key=lambda item: item[1], reverse=True)
    if numeric_counts and numeric_counts[0][1] > 0:
        return numeric_counts[0][0]
    raise ValueError("Could not find a trial balance amount column")


def _entity_tokens(entity: str) -> set[str]:
    tokens = set(_normalize(entity).split())
    compact = _compact(entity)
    tokens.add(compact)

    if "hk" in tokens or "hong" in tokens or "kong" in tokens:
        tokens.update({"hk", "hongkong", "hklimited"})
    if "uk" in tokens:
        tokens.update({"uk", "unitedkingdom", "uklimited"})
    if "europe" in tokens:
        tokens.update({"europe", "europesas", "sas"})
    if "technologies" in tokens:
        tokens.update({"technologies", "pltechnologies", "uae"})
    return {t for t in tokens if len(t) > 1 and t not in {"pl", "limited", "ltd"}}


def _infer_entity(
    path: Path,
    df: pd.DataFrame,
    balance_col: str,
    entities: list[str],
    source_name: str | None = None,
) -> tuple[str, float]:
    haystack = _compact(f"{source_name or path.name} {balance_col} {' '.join(map(str, df.columns))}")
    scores: list[tuple[float, str]] = []
    for entity in entities:
        score = 0.0
        for token in _entity_tokens(entity):
            if token and token in haystack:
                score += 3.0 if len(token) > 3 else 1.0
        scores.append((score, entity))

    scores.sort(reverse=True)
    best_score, best_entity = scores[0]
    if best_score <= 0:
        raise ValueError(f"Could not infer which Light entity '{path.name}' belongs to")
    confidence = best_score / max(best_score + (scores[1][0] if len(scores) > 1 else 0), 1.0)
    return best_entity, round(confidence, 3)


def _read_trial_balance(
    path: Path,
    entities: list[str],
    entity_override: str | None = None,
    source_name: str | None = None,
) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame]:
    df = _read_table(path)
    description_col = _find_column(df, ["Description", "Account", "Account name", "Account code"], contains=True)
    balance_col = _balance_column(df)
    entity, confidence = (
        (entity_override, 1.0)
        if entity_override
        else _infer_entity(path, df, balance_col, entities, source_name=source_name)
    )

    working = pd.DataFrame(index=df.index)
    working["entity"] = entity
    working["account_code"] = df[description_col].apply(_account_code)
    working["account_description"] = df[description_col].apply(_account_description)
    working["tb_balance"] = df[balance_col].apply(_amount)
    working = working[working["account_code"] != ""]

    grouped = (
        working.groupby(["entity", "account_code"], dropna=False)
        .agg(
            tb_balance=("tb_balance", "sum"),
            tb_rows=("tb_balance", "size"),
            account_description=("account_description", "first"),
        )
        .reset_index()
    )
    grouped["tb_source_file"] = path.name

    meta = {
        "source_file": path.name,
        "entity": entity,
        "confidence": confidence,
        "balance_column": str(balance_col),
        "rows": int(len(working)),
        "accounts": int(grouped["account_code"].nunique()),
        "total_balance": float(grouped["tb_balance"].sum()),
    }
    return grouped, meta, grouped.copy()


def _read_light_journals(paths: list[Path]) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    metas: list[dict[str, Any]] = []
    line_frames: list[pd.DataFrame] = []
    for path in paths:
        frame, meta, lines = _read_light_journal(path)
        frames.append(frame)
        metas.append(meta)
        line_frames.append(lines)

    if not frames:
        raise ValueError("At least one Light journal file is required")

    combined = pd.concat(frames, ignore_index=True)
    grouped = (
        combined.groupby(["entity", "account_code"], dropna=False)
        .agg(
            light_debit=("light_debit", "sum"),
            light_credit=("light_credit", "sum"),
            light_balance=("light_balance", "sum"),
            light_lines=("light_lines", "sum"),
            light_entries=("light_entries", "sum"),
            currency=("currency", lambda s: ", ".join(sorted({v for v in s if v and v.lower() != "nan"}))),
        )
        .reset_index()
    )

    entities = sorted({entity for meta in metas for entity in meta["entities"]})
    meta = {
        "rows": int(sum(meta["rows"] for meta in metas)),
        "entities": entities,
        "accounts": int(grouped["account_code"].nunique()),
        "total_debit": float(sum(meta["total_debit"] for meta in metas)),
        "total_credit": float(sum(meta["total_credit"] for meta in metas)),
        "source_files": [path.name for path in paths],
    }
    return grouped, meta, pd.concat(line_frames, ignore_index=True)


def classify_reconciliation_file(path: Path, source_name: str | None = None) -> dict[str, Any]:
    """Classify a file for the reconciliation batch flow."""
    display_name = source_name or path.name
    journal_score = 0.0
    tb_score = 0.0
    reasons: list[str] = []

    try:
        journal_df = _read_table(path, ["Entity", "Account code", "Debit", "Credit"])
        _find_column(journal_df, ["Entity", "Company", "Legal entity"])
        _find_column(journal_df, ["Account code", "Account", "GL account", "Ledger account"], contains=True)
        debit_col = _find_column(journal_df, ["Debit", "Debit amount"], contains=True)
        credit_col = _find_column(journal_df, ["Credit", "Credit amount"], contains=True)
        debit_total = journal_df[debit_col].apply(_amount).abs().sum()
        credit_total = journal_df[credit_col].apply(_amount).abs().sum()
        if len(journal_df) > 0 and (debit_total > 0 or credit_total > 0):
            journal_score = 0.95
            reasons.append("found Entity, Account code, Debit, and Credit columns")
    except Exception:
        pass

    try:
        tb_df = _read_table(path)
        description_col = _find_column(tb_df, ["Description", "Account", "Account name", "Account code"], contains=True)
        balance_col = _balance_column(tb_df)
        account_codes = tb_df[description_col].apply(_account_code)
        account_like_rows = account_codes.astype(str).str.contains(r"\d", regex=True).sum()
        balance_total = tb_df[balance_col].apply(_amount).abs().sum()
        if account_like_rows >= 3 and balance_total > 0:
            balance_key = _normalize(balance_col)
            tb_score = 0.9 if "balance" in balance_key else 0.7
            reasons.append(f"found account descriptions and balance column '{balance_col}'")
    except Exception:
        pass

    if journal_score >= 0.75 and journal_score >= tb_score:
        kind = "journal"
        confidence = journal_score
    elif tb_score >= 0.7:
        kind = "trial_balance"
        confidence = tb_score
    else:
        kind = "unknown"
        confidence = max(journal_score, tb_score)
        if not reasons:
            reasons.append("no recognizable Light journal or trial balance structure")

    return {
        "filename": display_name,
        "stored_filename": path.name,
        "kind": kind,
        "confidence": round(confidence, 3),
        "reason": "; ".join(reasons),
    }


def reconcile_light_journal_to_trial_balances(
    journal_path: Path | list[Path],
    trial_balance_paths: list[Path],
    *,
    tolerance: float = 0.01,
    entity_overrides: dict[str, str] | None = None,
    source_names: dict[str, str] | None = None,
) -> dict[str, Any]:
    if not trial_balance_paths:
        raise ValueError("At least one trial balance file is required")

    journal_paths = [journal_path] if isinstance(journal_path, Path) else journal_path
    journal_df, journal_meta, journal_lines_df = _read_light_journals(journal_paths)
    entities = journal_meta["entities"]
    if not entities:
        raise ValueError("No entities were found in the Light journal file")

    tb_frames: list[pd.DataFrame] = []
    tb_mappings: list[dict[str, Any]] = []
    tb_tabs: list[dict[str, Any]] = []
    overrides = entity_overrides or {}
    names = source_names or {}
    for tb_path in trial_balance_paths:
        tb_df, tb_meta, tb_tab_df = _read_trial_balance(
            tb_path,
            entities,
            overrides.get(tb_path.name) or overrides.get(str(tb_path)),
            source_name=names.get(tb_path.name) or names.get(str(tb_path)),
        )
        tb_frames.append(tb_df)
        tb_mappings.append(tb_meta)
        tb_tabs.append({
            "source_file": tb_path.name,
            "entity": tb_meta["entity"],
            "balance_column": tb_meta["balance_column"],
            "df": tb_tab_df,
        })

    trial_df = pd.concat(tb_frames, ignore_index=True)
    trial_df = (
        trial_df.groupby(["entity", "account_code"], dropna=False)
        .agg(
            tb_balance=("tb_balance", "sum"),
            tb_rows=("tb_rows", "sum"),
            account_description=("account_description", "first"),
            tb_source_file=("tb_source_file", lambda s: ", ".join(sorted(set(s)))),
        )
        .reset_index()
    )

    details = pd.merge(journal_df, trial_df, on=["entity", "account_code"], how="outer")
    details["light_debit"] = details["light_debit"].fillna(0.0)
    details["light_credit"] = details["light_credit"].fillna(0.0)
    details["light_balance"] = details["light_balance"].fillna(0.0)
    details["light_lines"] = details["light_lines"].fillna(0).astype(int)
    details["light_entries"] = details["light_entries"].fillna(0).astype(int)
    details["currency"] = details["currency"].fillna("")
    details["tb_balance"] = details["tb_balance"].fillna(0.0)
    details["tb_rows"] = details["tb_rows"].fillna(0).astype(int)
    details["account_description"] = details["account_description"].fillna("")
    details["tb_source_file"] = details["tb_source_file"].fillna("")
    details["variance"] = details["light_balance"] - details["tb_balance"]
    details["abs_variance"] = details["variance"].abs()

    def status(row: pd.Series) -> str:
        if row["abs_variance"] <= tolerance:
            return "matched"
        if row["light_lines"] == 0:
            return "missing_in_journal"
        if row["tb_rows"] == 0:
            return "missing_in_trial_balance"
        return "mismatch"

    details["status"] = details.apply(status, axis=1)
    status_rank = {
        "mismatch": 0,
        "missing_in_journal": 1,
        "missing_in_trial_balance": 2,
        "matched": 3,
    }
    details["_status_rank"] = details["status"].map(status_rank).fillna(9)
    details = details.sort_values(["_status_rank", "abs_variance"], ascending=[True, False]).reset_index(drop=True)
    details = details.drop(columns=["_status_rank"])

    entity_summary = (
        details.groupby("entity", dropna=False)
        .agg(
            accounts=("account_code", "count"),
            matched=("status", lambda s: int((s == "matched").sum())),
            mismatches=("status", lambda s: int((s == "mismatch").sum())),
            missing_in_journal=("status", lambda s: int((s == "missing_in_journal").sum())),
            missing_in_trial_balance=("status", lambda s: int((s == "missing_in_trial_balance").sum())),
            light_balance=("light_balance", "sum"),
            trial_balance=("tb_balance", "sum"),
            variance=("variance", "sum"),
            max_abs_variance=("abs_variance", "max"),
        )
        .reset_index()
    )

    summary = {
        "journal_file": ", ".join(path.name for path in journal_paths),
        "journal_files": [path.name for path in journal_paths],
        "trial_balance_files": [p.name for p in trial_balance_paths],
        "tolerance": tolerance,
        "entities": int(entity_summary["entity"].nunique()),
        "accounts": int(len(details)),
        "matched": int((details["status"] == "matched").sum()),
        "mismatches": int((details["status"] == "mismatch").sum()),
        "missing_in_journal": int((details["status"] == "missing_in_journal").sum()),
        "missing_in_trial_balance": int((details["status"] == "missing_in_trial_balance").sum()),
        "light_balance": float(details["light_balance"].sum()),
        "trial_balance": float(details["tb_balance"].sum()),
        "variance": float(details["variance"].sum()),
        "max_abs_variance": float(details["abs_variance"].max() if len(details) else 0.0),
    }

    columns = [
        "entity",
        "account_code",
        "account_description",
        "currency",
        "light_debit",
        "light_credit",
        "light_balance",
        "trial_balance",
        "variance",
        "status",
        "light_lines",
        "light_entries",
        "tb_rows",
        "tb_source_file",
    ]
    details = details.rename(columns={"tb_balance": "trial_balance"})
    details = details[columns + ["abs_variance"]]

    return {
        "summary": summary,
        "entity_summary": entity_summary.to_dict(orient="records"),
        "account_details": details.drop(columns=["abs_variance"]).to_dict(orient="records"),
        "journal_meta": journal_meta,
        "trial_balance_mappings": tb_mappings,
        "details_df": details,
        "journal_lines_df": journal_lines_df,
        "tb_tabs": tb_tabs,
    }


def _clean_cell_value(value: Any) -> Any:
    if isinstance(value, list | tuple | set):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join("-" if char in _INVALID_SHEET_CHARS else char for char in name).strip()
    return (cleaned or "Sheet")[:31]


def _unique_sheet_name(wb: openpyxl.Workbook, desired: str) -> str:
    name = _sanitize_sheet_name(desired)
    if name not in wb.sheetnames:
        return name
    base = name[:28]
    index = 2
    while f"{base}({index})" in wb.sheetnames:
        index += 1
    return f"{base}({index})"


def _quote_sheet(sheet_name: str) -> str:
    return f"'{sheet_name.replace(chr(39), chr(39) + chr(39))}'"


def _write_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([_clean_cell_value(value) for value in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = PatternFill("solid", fgColor="F3F4F6")


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 46)


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    rows = df.astype(object).values.tolist()
    _write_rows(ws, list(df.columns), rows)


def _journal_export_df(result: dict[str, Any]) -> pd.DataFrame:
    journal_df = result["journal_lines_df"].copy()
    columns = [
        "source_file",
        "entity",
        "account_code",
        "entry_id",
        "currency",
        "debit",
        "credit",
        "light_balance",
    ]
    original_columns = [col for col in journal_df.columns if str(col).startswith("Original - ")]
    return journal_df[columns + original_columns].rename(columns={
        "source_file": "Source file",
        "entity": "Entity",
        "account_code": "Account code",
        "entry_id": "Entry ID",
        "currency": "Currency",
        "debit": "Debit",
        "credit": "Credit",
        "light_balance": "Journal balance",
    })


def _tb_sheet_df(tb_tab: dict[str, Any], details_df: pd.DataFrame) -> pd.DataFrame:
    tb_df = tb_tab["df"].copy()
    entity = tb_tab["entity"]
    tb_df["entity"] = entity

    journal_only = details_df[
        (details_df["entity"] == entity)
        & (details_df["status"] == "missing_in_trial_balance")
    ][["entity", "account_code", "account_description"]].copy()
    if not journal_only.empty:
        journal_only["tb_balance"] = 0.0
        journal_only["tb_rows"] = 0
        journal_only["tb_source_file"] = tb_tab["source_file"]
        tb_df = pd.concat([tb_df, journal_only], ignore_index=True)

    tb_df = tb_df.drop_duplicates(["entity", "account_code"], keep="first")
    return tb_df[[
        "entity",
        "account_code",
        "account_description",
        "tb_balance",
        "tb_rows",
    ]].sort_values(["entity", "account_code"]).rename(columns={
        "entity": "Entity",
        "account_code": "Account code",
        "account_description": "Account description",
        "tb_balance": "TB balance",
        "tb_rows": "TB rows",
    })


def write_reconciliation_workbook(result: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df = pd.DataFrame([result["summary"]])
    entity_df = pd.DataFrame(result["entity_summary"])
    details_df = result["details_df"].drop(columns=["abs_variance"], errors="ignore")
    mapping_df = pd.DataFrame(result["trial_balance_mappings"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Entries"
    journal_export = _journal_export_df(result)
    _write_dataframe(ws, journal_export)

    journal_sheet_ref = _quote_sheet(ws.title)
    journal_last_row = max(ws.max_row, 2)
    entity_range = f"{journal_sheet_ref}!$B$2:$B${journal_last_row}"
    account_range = f"{journal_sheet_ref}!$C$2:$C${journal_last_row}"
    debit_range = f"{journal_sheet_ref}!$F$2:$F${journal_last_row}"
    credit_range = f"{journal_sheet_ref}!$G$2:$G${journal_last_row}"

    for tb_tab in result.get("tb_tabs", []):
        source_name = Path(str(tb_tab["source_file"])).stem
        tb_ws = wb.create_sheet(_unique_sheet_name(wb, source_name))
        tb_df = _tb_sheet_df(tb_tab, details_df)
        headers = [
            "Entity",
            "Account code",
            "Account description",
            "TB balance",
            "TB rows",
            "Journal debit",
            "Journal credit",
            "Journal balance",
            "Variance",
            "Status",
        ]
        _write_rows(tb_ws, headers, tb_df.values.tolist())
        tb_ws["L1"] = "Tolerance"
        tb_ws["M1"] = result["summary"]["tolerance"]
        tb_ws["M1"].number_format = "#,##0.00"
        for row_idx in range(2, tb_ws.max_row + 1):
            tb_ws[f"F{row_idx}"] = f"=SUMIFS({debit_range},{entity_range},$A{row_idx},{account_range},$B{row_idx})"
            tb_ws[f"G{row_idx}"] = f"=SUMIFS({credit_range},{entity_range},$A{row_idx},{account_range},$B{row_idx})"
            tb_ws[f"H{row_idx}"] = f"=F{row_idx}-G{row_idx}"
            tb_ws[f"I{row_idx}"] = f"=H{row_idx}-D{row_idx}"
            tb_ws[f"J{row_idx}"] = (
                f'=IF(ABS(I{row_idx})<=$M$1,"matched",'
                f'IF(AND(F{row_idx}=0,G{row_idx}=0),"missing_in_journal",'
                f'IF(E{row_idx}=0,"missing_in_trial_balance","mismatch")))'
            )

    for sheet_name, df in [
        ("Summary", summary_df),
        ("Entity Summary", entity_df),
        ("Account Detail", details_df),
        ("Source Mapping", mapping_df),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_dataframe(ws, df)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    for sheet in wb.worksheets:
        _autosize(sheet)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header and "Account code" in str(header):
                letter = get_column_letter(col)
                for cell in sheet[letter]:
                    cell.number_format = "@"

    wb.save(output_path)

    return output_path
