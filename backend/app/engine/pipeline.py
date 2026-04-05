from enum import Enum
from pathlib import Path

import pandas as pd

from app.engine.registry import get_transform, get_validator
from app.engine.file_transforms import get_file_transform
from app.services.file_service import read_file_as_dataframe
from app.services.playbook_service import PlaybookConfig


class PipelineStage(str, Enum):
    INGEST = "ingest"
    DETECT = "detect"
    MAP = "map"
    TRANSFORM = "transform"
    VALIDATE = "validate"
    EXPORT = "export"


class PipelineResult:
    def __init__(self):
        self.stage: PipelineStage | None = None
        self.dataframes: dict[str, pd.DataFrame] = {}  # file_id -> df
        self.transformed: dict[str, pd.DataFrame] = {}  # file_id -> transformed df
        self.validation_issues: list[dict] = []
        self.output_path: Path | None = None
        self.error: str | None = None


def is_file_transform(file_type_config: dict) -> bool:
    return file_type_config.get("transform_type") == "file"


def is_rule_pipeline(file_type_config: dict) -> bool:
    return file_type_config.get("transform_type") == "rule_pipeline"


def run_rule_pipeline(file_path: Path, rules: list[dict], limit: int | None = None) -> pd.DataFrame:
    """Execute an ordered list of composable rules as a pipeline."""
    from app.engine.rules import get_rule_executor

    sheets = pd.read_excel(file_path, sheet_name=None, header=None, engine="openpyxl")
    df = pd.DataFrame()

    for rule in rules:
        if not rule.get("enabled", True):
            continue

        rule_type = rule["type"]
        config = rule.get("config", {})

        executor = get_rule_executor(rule_type)
        df = executor.execute(df, sheets, config)

        # For preview mode, truncate early for performance
        if limit and len(df) > limit * 3:
            df = df.head(limit * 3)

    if limit:
        df = df.head(limit)

    return df


def run_file_transform(file_path: Path, file_type_config: dict) -> pd.DataFrame:
    """Read all sheets from the workbook and run a registered file-level transform."""
    handler_name = file_type_config.get("handler")
    if not handler_name:
        raise ValueError("file transform config missing 'handler'")

    sheets = pd.read_excel(file_path, sheet_name=None, header=None, engine="openpyxl")
    handler = get_file_transform(handler_name)
    return handler(sheets, file_type_config.get("params", {}))


def run_ingest(file_path: Path, file_type_config: dict) -> pd.DataFrame:
    sheet = file_type_config.get("source_sheet", 0)
    header_row = file_type_config.get("header_row", 0)
    skip_footer = file_type_config.get("skip_footer_rows", 0)
    return read_file_as_dataframe(file_path, sheet=sheet, header_row=header_row, skip_footer=skip_footer)


def run_transform(df: pd.DataFrame, file_type_config: dict) -> pd.DataFrame:
    mappings = file_type_config.get("mappings", [])
    result = pd.DataFrame(index=df.index)

    for mapping in mappings:
        source_col = mapping.get("source")
        target_col = mapping["target"]
        transform_name = mapping.get("transform", "passthrough")
        params = mapping.get("params", {})

        if source_col is None:
            # Computed/static column - pass an empty series
            source_series = pd.Series([""] * len(df), index=df.index)
        elif source_col in df.columns:
            source_series = df[source_col].copy()
        else:
            # Source column not found - create empty
            source_series = pd.Series([""] * len(df), index=df.index)

        # Support chained transforms
        transforms_list = mapping.get("transforms")
        if transforms_list:
            series = source_series
            for t in transforms_list:
                func = get_transform(t["name"])
                series = func(series, t.get("params", {}))
            result[target_col] = series
        else:
            func = get_transform(transform_name)
            result[target_col] = func(source_series, params)

    return result


def run_validate(df: pd.DataFrame, file_type_config: dict) -> list[dict]:
    validations = file_type_config.get("validations", [])
    all_issues = []

    for val_config in validations:
        validator_name = val_config["validator"]
        severity = val_config.get("severity", "warning")
        params = val_config.get("params", {})

        func = get_validator(validator_name)
        issues = func(df, params)

        # Override severity from config
        for issue in issues:
            issue["severity"] = severity
        all_issues.extend(issues)

    return all_issues


_INVALID_SHEET_CHARS = set(r':\/?*[]')


def _sanitize_sheet_name(name: str) -> str:
    """Make a sheet name Excel-safe: strip forbidden chars and cap at 31 chars."""
    cleaned = ''.join('-' if c in _INVALID_SHEET_CHARS else c for c in name).strip()
    if not cleaned:
        cleaned = 'working file - output'
    return cleaned[:31]


def _unique_sheet_name(wb, desired: str) -> str:
    """Return a sheet name not already in the workbook, respecting the 31-char limit."""
    if desired not in wb.sheetnames:
        return desired
    base = desired[:28]  # leave room for suffix
    i = 2
    while f"{base}({i})" in wb.sheetnames:
        i += 1
    return f"{base}({i})"


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def run_export(
    df: pd.DataFrame,
    output_path: Path,
    template_path: Path | None = None,
    source_file_path: Path | None = None,
    working_sheet_topic: str | None = None,
) -> Path:
    """Write the transformed dataframe to output_path.

    When source_file_path is provided, the original workbook is preserved in full
    (all sheets untouched) and the transformed result is appended as a new sheet
    named "working file - <topic>". This gives the user an audit trail: the
    original source and the migration result in a single file.
    """
    import openpyxl

    working_sheet_name = _sanitize_sheet_name(
        f"working file - {working_sheet_topic}" if working_sheet_topic else "working file - output"
    )

    if source_file_path and Path(source_file_path).exists():
        # Preserve original workbook; append transformed result as a new sheet.
        wb = openpyxl.load_workbook(source_file_path)
        sheet_name = _unique_sheet_name(wb, working_sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        _write_df_to_sheet(ws, df)
        wb.save(output_path)
    elif template_path and template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        _write_df_to_sheet(ws, df)
        wb.save(output_path)
    else:
        # No source and no template - create fresh workbook
        df.to_excel(output_path, index=False, engine="openpyxl")

    return output_path
